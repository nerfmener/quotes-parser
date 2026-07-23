import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import os


class ExcelAutomation:
    """Класс для автоматизации работы с Excel"""

    def __init__(self, output_dir="output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def merge_multiple_files(self, file_paths, output_name="merged.xlsx"):
        """
        tОбъединяет несколько Excel-файлов в один

        Args:
            file_paths: список путей к файлам
            output_name: имя выходного файла
        """
        all_data = []

        for file_path in file_paths:
            df = pd.read_excel(file_path)
            df['source_file'] = os.path.basename(file_path)
            all_data.append(df)

        merged_df = pd.concat(all_data, ignore_index=True)
        output_path = os.path.join(self.output_dir, output_name)
        merged_df.to_excel(output_path, index=False)

        print(f"✅ Объединено {len(file_paths)} файлов")
        print(f"📊 Всего строк: {len(merged_df)}")
        print(f"💾 Сохранено: {output_path}")

        return output_path

    def generate_report(self, data, output_name="report.xlsx"):
        """
        Генерирует красивый отчёт с графиками

        Args:
            data: DataFrame с данными
            output_name: имя выходного файла
        """
        output_path = os.path.join(self.output_dir, output_name)

        # Создаём Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Основные данные
            data.to_excel(writer, sheet_name='Данные', index=False)

            # Сводная таблица
            if len(data.columns) >= 2:
                pivot = data.groupby(data.columns[0]).size().reset_index(name='Количество')
                pivot.to_excel(writer, sheet_name='Сводка', index=False)

        # Форматируем
        self._format_excel(output_path)

        print(f"✅ Отчёт создан: {output_path}")
        return output_path

    def _format_excel(self, file_path):
        """Форматирует Excel-файл (цвета, шрифты, границы)"""
        wb = openpyxl.load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Заголовки
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Границы
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        wb.save(file_path)

    def process_large_file(self, file_path, chunk_size=10000):
        """
        Обрабатывает большой Excel-файл по частям

        Args:
            file_path: путь к файлу
            chunk_size: размер чанка
        """
        total_rows = 0

        for chunk in pd.read_excel(file_path, chunksize=chunk_size):
            # Обработка чанка
            chunk = chunk.drop_duplicates()
            chunk = chunk.dropna(how='all')
            total_rows += len(chunk)

            print(f"📦 Обработано: {total_rows} строк")

        print(f"✅ Всего обработано: {total_rows} строк")
        return total_rows
if __name__ == "__main__":
    auto = ExcelAutomation()
    
    # Пример 1: Создание тестовых данных
    sample_data = pd.DataFrame({
        'Товар': ['Ноутбук', 'Телефон', 'Планшет', 'Наушники', 'Мышь'],
        'Цена': [50000, 30000, 25000, 5000, 1500],
        'Количество': [10, 25, 15, 50, 100],
        'Категория': ['Электроника', 'Электроника', 'Электроника', 'Аксессуары', 'Аксессуары']
    })
    
    # Пример 2: Генерация отчёта
    auto.generate_report(sample_data, "sales_report.xlsx")
    
    print("\n📊 Примеры использования:")
    print("1. Объединение файлов: auto.merge_multiple_files(['file1.xlsx', 'file2.xlsx'])")
    print("2. Генерация отчёта: auto.generate_report(dataframe, 'report.xlsx')")
    print("3. Обработка больших файлов: auto.process_large_file('big_file.xlsx')")    